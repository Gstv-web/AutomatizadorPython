from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font


class ReportWriter:

    SENT_SHEET_NAME = "RELATÓRIO"
    RECEIVED_SHEET_NAME = "RESPOSTAS"
    INTERACTION_SHEET_NAME = "INTERAÇÃO"
    ACTIVATION_SHEET_NAME = "Ativação"
    CALL_INTERACTION_SHEET_NAME = "Interação"


    def __init__(self, file_name):
        self.workbook = None
        self.file_name = None
        self.sent_sheet = None
        self.received_sheet = None
        self.interaction_sheet = None
        self.activation_sheet = None
        self.call_interaction_sheet = None
        self.file_name = file_name
        self.init_writer()

    def init_writer(self):
        self.workbook = Workbook()
        self.sent_sheet = self.workbook.active
        self.sent_sheet.title = ReportWriter.SENT_SHEET_NAME
        # self.sent_sheet = self.workbook.create_sheet(ReportWriter.SENT_SHEET_NAME)
        self.received_sheet = self.workbook.create_sheet(ReportWriter.RECEIVED_SHEET_NAME)
        self.insert_sent_header()
        self.insert_received_header()
        self.set_styles()

    def create_interaction_sheet(self):
        self.interaction_sheet = self.workbook.create_sheet(ReportWriter.INTERACTION_SHEET_NAME)

    def create_activation_sheet(self):
        self.activation_sheet = self.workbook.create_sheet(ReportWriter.ACTIVATION_SHEET_NAME)
    
    def create_call_interaction_sheet(self):
        self.call_interaction_sheet = self.workbook.create_sheet(ReportWriter.CALL_INTERACTION_SHEET_NAME)

    def set_sent_styles(self):
        self.sent_sheet.column_dimensions['A'].width = 17
        self.sent_sheet.column_dimensions['B'].width = 14
        self.sent_sheet.column_dimensions['C'].width = 14
        self.sent_sheet['A1'].font = Font(bold=True)
        self.sent_sheet['B1'].font = Font(bold=True)
        self.sent_sheet['C1'].font = Font(bold=True)
        for rows in self.sent_sheet.iter_rows(min_row=1):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        for col in self.sent_sheet.iter_cols(min_row=1, min_col=1, max_col=3):
            for cell in col:
                cell.border = thin_border

    def set_received_styles(self):
        self.received_sheet.column_dimensions['A'].width = 17
        self.received_sheet.column_dimensions['B'].width = 300
        self.received_sheet['A1'].font = Font(bold=True)
        self.received_sheet['B1'].font = Font(bold=True)
        for rows in self.received_sheet.iter_rows(min_row=1):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        for col in self.received_sheet.iter_cols(min_row=1, min_col=1, max_col=2):
            for cell in col:
                cell.border = thin_border

    # def set_interaction_styles(self):
    #     for rows in self.interaction_sheet.iter_cols(min_row=1, min_col=1):
    #         for cell in rows:
                
    #             cell.alignment = Alignment(horizontal='center')
    #     thin_border = Border(left=Side(style='thin'),
    #                         right=Side(style='thin'),
    #                         top=Side(style='thin'),
    #                         bottom=Side(style='thin'))
    #     for col in self.interaction_sheet.iter_cols(min_row=1, min_col=1, max_col=2):
    #         for cell in col:
    #             cell.border = thin_border


                

    def fix_number(self):
        if '55' in number and len(number) >= 12:
            number = number.lstrip('55')
        if "(" ")" " " in number:
            number = number.replace("(", "").replace(")", "").replace(" ", "")
    
    def set_styles(self):
        self.set_received_styles()
        self.set_sent_styles()

    def write_sent_message(self, number, sent_date, status):
 
        self.sent_sheet.append([number, sent_date, status])

    def write_received_message(self, number, received_text):
        
        self.received_sheet.append([number, received_text])
    
    def write_interaction_message(self, values):
        # self.interaction_sheet = self.workbook.create_sheet(ReportWriter.INTERACTION_SHEET_NAME)
        self.interaction_sheet.append(values)
        self.interaction_sheet['A1'] = 'TELEFONE'
        for row in self.interaction_sheet.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in row:
                cell.font = Font(bold=True)
        for rows in self.interaction_sheet.iter_cols(min_row=1, min_col=1):
            for cell in rows:
                cell.alignment = Alignment(horizontal='center')
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        for col in self.interaction_sheet.iter_cols(min_row=1, min_col=1):
            for cell in col:
                cell.border = thin_border

    def write_activation_message(self, values):
        self.activation_sheet.append(values)

    def write_call_interaction_message(self, values):
        self.call_interaction_sheet.append(values)

    def insert_sent_header(self):
        self.sent_sheet.insert_rows(0)
        self.sent_sheet['A1'] = 'TELEFONE'
        self.sent_sheet['B1'] = 'DATA'
        self.sent_sheet['C1'] = 'STATUS'

    def insert_received_header(self):  
        self.received_sheet.insert_rows(0)
        self.received_sheet['A1'] = 'TELEFONE'
        self.received_sheet['B1'] = 'MENSAGEM'

    

    # def remove_55_from_numbers(self):
    #     for col in self.sent_sheet.iter_cols(min_row=2, min_col=1, max_col=1):
    #         for cell in col:
    #             new_number = str(cell.value)
    #             if '55' in cell.value:
    #                 new_number = new_number.lstrip('55')
    #                 cell.value = new_number
    #     for col in self.received_sheet.iter_cols(min_row=1, min_col=1, max_col=1):
    #         for cell in col:
    #             new_number = str(cell.value)
    #             if '55' in cell.value:
    #                 new_number = new_number.lstrip('55')
    #                 cell.value = new_number 


    def edit_date(self):
        for col in self.sent_sheet.iter_cols(min_row=2, min_col=2, max_col=2):
            for cell in col:
                celula = str(cell.value)  # aqui eu acessei o valor da célula, agora eu posso alterar o valor
                dateSet = celula.split()
                justDate = dateSet[0]
                if '"' in justDate:
                    justDate = justDate.replace('"', '')
                justDate.replace("\"", "") # Aqui eu acesso apenas a data
                cell.value = justDate

    # def edit_status(self):
    #     new_status = 'Enviado'
    #     for col in self.sent_sheet.iter_cols(min_row=2, min_col=3, max_col=3):
    #         for cell in col:
    #             cell.value = new_status

    def remove_blacklist_numbers(self):
        blacklist = ['71983155252', '62993984240', '11986884500',
                     '21997610334', '62985553937', '11997399665',
                     '47999905666', '11967243506', '11939390655',
                     '89988169173', '11995458123', '41996166105',
                     '11996384589', '71991810512', '47999474163',
                     '11977826644', '11997965365', '11980831080',
                     '11982026091', '11995556294', '71991770338',
                     '11996869475']
            
        for col in self.sent_sheet.iter_cols(min_row=2, min_col=1, max_col=1):
            for cell in col:
                if cell.value in blacklist:
                    self.sent_sheet.delete_rows(cell.row)
        

    def remove_blank_messages(self):
        for col in self.received_sheet.iter_cols(min_row=2, min_col=2, max_col=2):
            for cell in col:
                if cell.value == '\n' or cell.value == None:
                    self.received_sheet.delete_rows(cell.row)

    def save(self):
        self.workbook.save(self.file_name)