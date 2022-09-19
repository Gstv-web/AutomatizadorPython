from report_converter import ReportConverterInterface
from report import Report
from openpyxl import load_workbook


class LigdataConverter(ReportConverterInterface):

    def convert_wpp_report(self, report: Report):
        self.cast_sent_report(report.sent_file)
        self.cast_received_report(report.received_file)
        self.writer.set_styles()
        self.writer.remove_blacklist_numbers()
        self.writer.remove_blank_messages()
        self.writer.save()

    def cast_sent_report(self, sent_file_name):
        workbook = load_workbook(sent_file_name)
        worksheet = workbook['Envio']
        worksheet.delete_rows(1)
        date = worksheet['A']
        phone = worksheet['B']
        status = worksheet['C']

        date_val = [cell.value for cell in date]
        phone_val = [cell.value for cell in phone]
        status_val = [cell.value for cell in status]

        for p, d, s in zip(phone_val, date_val, status_val):
            self.writer.write_sent_message(self.fix_number(str(p)), self.fix_date(str(d)), s)

    def cast_received_report(self, received_file_name):
        workbook = load_workbook(received_file_name)
        worksheet = workbook['Respostas']
        worksheet.delete_rows(1)
        phone = worksheet['A']
        message = worksheet['B']
        
        phone_val = [cell.value for cell in phone]
        message_val = [cell.value for cell in message]

        for p, m in zip(phone_val, message_val):
            self.writer.write_received_message(self.fix_number(str(p)), m)

    
