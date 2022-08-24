from report_converter import ReportConverterInterface
from report import Report
from openpyxl import load_workbook


class AtmosferaConverter(ReportConverterInterface):

    def convert_report(self, report:Report):
        self.cast_sent_report(report.sent_file)
        self.cast_received_report(report.received_file)
        self.writer.set_styles()
        self.writer.remove_blacklist_numbers()
        self.writer.remove_blank_messages()
        self.writer.save()

    def cast_sent_report(self, sent_file_name):
        workbook = load_workbook(sent_file_name) 
        worksheet = workbook['RELATÓRIO']
        
        phone = worksheet['B']
        date = worksheet['C']
        status = worksheet['H']


        phone_val = [cell.value for cell in phone if cell.value != None]
        date_val = [cell.value for cell in date if cell.value != None]
        status_val = [cell.value for cell in status if cell.value != None]
       
        for p, d, s in zip(phone_val, date_val, status_val):
            if s == "Yes":
                self.writer.write_sent_message(self.fix_number(str(p)), self.fix_date(str(d)), self.fix_status(s))



    def cast_received_report(self, received_file_name):
        workbook = load_workbook(received_file_name)
        worksheet = workbook['RESPOSTAS']

        phone = worksheet['B']
        message = worksheet['C']

        phone_val = [cell.value for cell in phone if cell.value != None and cell.value != 'Contatos']
        message_val = [cell.value for cell in message if cell.value != None and cell.value != 'Respostas']

        next(zip(phone_val, message_val))
        for p, m in zip(phone_val, message_val):
            self.writer.write_received_message(self.fix_number(str(p)), m)